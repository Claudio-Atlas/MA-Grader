# orchestrator/phase1_grade_all.py

import os
from openpyxl import load_workbook
from typing import Dict, Any, Optional

from utilities.logger import get_logger

from graders.income_analysis.grade_income_analysis import grade_income_analysis
from writers.write_income_analysis_scores import write_income_analysis_scores

# ---- Unit Conversions V2 imports ----
from graders.unit_conversions.unit_conversions_checker_v2 import grade_unit_conversions_tab_v2
from writers.unit_conversions_writer_v2 import write_unit_conversions_scores_v2
# -----------------------------------

# ---- Currency Conversion V2 imports ----
from graders.currency_conversion.grade_currency_conversion_tab_v2 import grade_currency_conversion_tab_v2
from writers.write_currency_conversion_results_v2 import write_currency_conversion_results_v2
# ---------------------------------------

# ---- Sheet Validation ----
from utilities.validate_submission import validate_required_sheets, log_missing_sheets
# --------------------------


def phase1_grade_all_students(
    submissions_path: str,
    graded_output_path: str,
    pipeline_state: Optional[Dict[str, Any]] = None
) -> None:
    """
    Grades the formula-based parts of every student's MA1 workbook.
    (Chart export and insertion happen in later phases.)
    
    Args:
        submissions_path: Path to folder containing student submission files
        graded_output_path: Path to folder containing grading sheet templates
        pipeline_state: Optional dict for cancellation checking (from server.py)
    """
    logger = get_logger()
    logger.info("")
    logger.info("=" * 60)
    logger.info("PHASE 1 - Grading all students...")
    logger.info("=" * 60)

    student_files = [f for f in os.listdir(submissions_path) if f.endswith(".xlsx")]
    total_students = len(student_files)
    logger.info(f"Found {total_students} student submissions to grade")

    graded_count = 0
    error_count = 0
    skipped_count = 0

    for idx, filename in enumerate(student_files, 1):
        # Check for cancellation request
        if pipeline_state and pipeline_state.get("cancel_requested"):
            logger.warning(f"Pipeline cancelled by user after grading {graded_count} students")
            break

        student_name = filename.replace("_MA1.xlsx", "")
        submission_file = os.path.join(submissions_path, filename)
        grading_file = os.path.join(graded_output_path, f"{student_name}_MA1_Grade.xlsx")

        logger.info(f"[{idx}/{total_students}] Processing: {student_name}")

        student_wb = None
        grading_wb = None
        
        try:
            logger.debug(f"  Loading submission: {submission_file}")
            student_wb = load_workbook(submission_file, data_only=False)
            
            logger.debug(f"  Loading grading sheet: {grading_file}")
            grading_wb = load_workbook(grading_file)

            ws_grading = grading_wb["Grading Sheet"]

            # --- Validate required sheets (case-insensitive) ---
            is_valid, sheet_map, missing_sheets = validate_required_sheets(student_wb)
            
            if missing_sheets:
                log_missing_sheets(student_name, missing_sheets)

            # -----------------------------
            # INCOME ANALYSIS
            # -----------------------------
            if sheet_map.get("Income Analysis"):
                try:
                    logger.debug(f"  Grading Income Analysis...")
                    ws_income = student_wb[sheet_map["Income Analysis"]]
                    ia_results = grade_income_analysis(ws_income)
                    write_income_analysis_scores(ws_grading, ia_results)
                    logger.debug(f"  Income Analysis complete")
                except Exception as e:
                    logger.warning(f"  Income Analysis error for {student_name}: {e}")
            else:
                logger.info(f"  Skipping Income Analysis (sheet missing)")
                skipped_count += 1

            # -----------------------------
            # UNIT CONVERSIONS - V2 ONLY
            # -----------------------------
            if sheet_map.get("Unit Conversions"):
                try:
                    logger.debug(f"  Grading Unit Conversions...")
                    ws_unit = student_wb[sheet_map["Unit Conversions"]]
                    uc_results = grade_unit_conversions_tab_v2(ws_unit)
                    write_unit_conversions_scores_v2(ws_grading, uc_results)
                    logger.debug(f"  Unit Conversions complete")
                except Exception as e:
                    logger.warning(f"  Unit Conversions error for {student_name}: {e}")
            else:
                logger.info(f"  Skipping Unit Conversions (sheet missing)")
                skipped_count += 1

            # -----------------------------
            # CURRENCY CONVERSION - V2 ONLY
            # -----------------------------
            if sheet_map.get("Currency Conversion"):
                try:
                    logger.debug(f"  Grading Currency Conversion...")
                    ws_currency = student_wb[sheet_map["Currency Conversion"]]
                    cc_results = grade_currency_conversion_tab_v2(ws_currency, student_name)
                    write_currency_conversion_results_v2(ws_grading, cc_results)
                    logger.debug(f"  Currency Conversion complete")
                except Exception as e:
                    logger.warning(f"  Currency Conversion error for {student_name}: {e}")
            else:
                logger.info(f"  Skipping Currency Conversion (sheet missing)")
                skipped_count += 1

            grading_wb.save(grading_file)
            logger.info(f"  ✓ Graded: {student_name}")
            graded_count += 1

        except Exception as e:
            logger.error(f"  ✗ Error grading {student_name}: {e}")
            error_count += 1
        
        finally:
            # Ensure workbooks are always closed to prevent file locks and memory leaks
            if student_wb is not None:
                student_wb.close()
            if grading_wb is not None:
                grading_wb.close()

    # Summary
    logger.info("")
    logger.info("-" * 40)
    logger.info(f"Phase 1 Summary: {graded_count} graded, {error_count} errors, {skipped_count} sheets skipped")
    logger.info("-" * 40)
