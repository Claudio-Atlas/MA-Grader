# orchestrator/grade_single.py

import os
from openpyxl import load_workbook

from utilities.logger import get_logger

from graders.income_analysis.grade_income_analysis import grade_income_analysis
from writers.write_income_analysis_scores import write_income_analysis_scores

# ---- Unit Conversions V2 ----
from graders.unit_conversions.unit_conversions_checker_v2 import grade_unit_conversions_tab_v2
from writers.unit_conversions_writer_v2 import write_unit_conversions_scores_v2
# -----------------------------

# ---- Currency Conversion V2 ----
from graders.currency_conversion.grade_currency_conversion_tab_v2 import grade_currency_conversion_tab_v2
from writers.write_currency_conversion_results_v2 import write_currency_conversion_results_v2
# --------------------------------

# ---- Sheet Validation ----
from utilities.validate_submission import validate_required_sheets, get_sheet_safe, log_missing_sheets
# --------------------------


def grade_single_file(submission_path: str, graded_output_folder: str) -> dict:
    """
    Grade a single student's MA1 workbook.

    Assumes:
      - submission_path points to:
            student_submissions/<course_label>/First_Last_MA1.xlsx
      - A matching grading sheet already exists:
            graded_output/<course_label>/First_Last_MA1_Grade.xlsx
    
    Args:
        submission_path: Path to the student's submission file
        graded_output_folder: Path to folder containing grading sheets
        
    Returns:
        dict: Results dictionary containing scores for each section
    """
    logger = get_logger()

    if not os.path.exists(submission_path):
        raise FileNotFoundError(f"Submission file not found: {submission_path}")

    if not os.path.isdir(graded_output_folder):
        raise FileNotFoundError(f"Graded output folder not found: {graded_output_folder}")

    # --- Infer grading sheet filename ---
    submission_name = os.path.basename(submission_path)
    base_name, ext = os.path.splitext(submission_name)
    grading_filename = f"{base_name}_Grade{ext}"
    grading_path = os.path.join(graded_output_folder, grading_filename)

    if not os.path.exists(grading_path):
        raise FileNotFoundError(f"Grading sheet not found: {grading_path}")

    # --- Load workbooks ---
    student_wb = None
    grading_wb = None
    
    try:
        logger.debug(f"Loading submission: {submission_path}")
        student_wb = load_workbook(submission_path, data_only=False)
        
        logger.debug(f"Loading grading sheet: {grading_path}")
        grading_wb = load_workbook(grading_path)

        # --- Grading sheet ---
        try:
            grading_ws = grading_wb["Grading Sheet"]
        except KeyError:
            raise KeyError("The 'Grading Sheet' tab was not found in the grading template workbook.")

        results_out = {}
        student_name = base_name.replace("_MA1", "")  # "First_Last"

        logger.info(f"Grading single file: {student_name}")

        # --- Validate required sheets (case-insensitive) ---
        is_valid, sheet_map, missing_sheets = validate_required_sheets(student_wb)
        
        if missing_sheets:
            log_missing_sheets(student_name, missing_sheets)
            results_out["missing_sheets"] = missing_sheets

        # ------------------------------
        # INCOME ANALYSIS
        # ------------------------------
        if sheet_map.get("Income Analysis"):
            try:
                logger.debug("  Grading Income Analysis...")
                ws_income = student_wb[sheet_map["Income Analysis"]]
                ia_results = grade_income_analysis(ws_income)
                write_income_analysis_scores(grading_ws, ia_results)
                results_out["income_analysis"] = ia_results
                logger.debug("  Income Analysis complete")
            except Exception as e:
                logger.warning(f"  Income Analysis error: {e}")
        else:
            logger.info(f"  Skipping Income Analysis (sheet missing)")

        # ------------------------------
        # UNIT CONVERSIONS - V2
        # ------------------------------
        if sheet_map.get("Unit Conversions"):
            try:
                logger.debug("  Grading Unit Conversions...")
                ws_unit = student_wb[sheet_map["Unit Conversions"]]
                uc_results = grade_unit_conversions_tab_v2(ws_unit)
                write_unit_conversions_scores_v2(grading_ws, uc_results)
                results_out["unit_conversions_v2"] = uc_results
                logger.debug("  Unit Conversions complete")
            except Exception as e:
                logger.warning(f"  Unit Conversions error: {e}")
        else:
            logger.info(f"  Skipping Unit Conversions (sheet missing)")

        # ------------------------------
        # CURRENCY CONVERSION - V2
        # ------------------------------
        if sheet_map.get("Currency Conversion"):
            try:
                logger.debug("  Grading Currency Conversion...")
                ws_currency = student_wb[sheet_map["Currency Conversion"]]
                cc_results = grade_currency_conversion_tab_v2(ws_currency, student_name)
                write_currency_conversion_results_v2(grading_ws, cc_results)
                results_out["currency_conversion_v2"] = cc_results
                logger.debug("  Currency Conversion complete")
            except Exception as e:
                logger.warning(f"  Currency Conversion error: {e}")
        else:
            logger.info(f"  Skipping Currency Conversion (sheet missing)")

        # Save the updated grading workbook
        grading_wb.save(grading_path)
        logger.info(f"  âœ“ Graded: {student_name}")

        return results_out
    
    finally:
        # Ensure workbooks are always closed to prevent file locks and memory leaks
        if student_wb is not None:
            student_wb.close()
        if grading_wb is not None:
            grading_wb.close()
