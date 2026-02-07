"""
phase1_grade_all_ma3.py — MA3 Grading orchestrator

Purpose: Grades the formula-based parts of every student's MA3 workbook.
         Handles Analysis and Visualization tabs.

Author: Clayton Ragsdale
"""

import os
from openpyxl import load_workbook
from typing import Dict, Any, Optional

from utilities.logger import get_logger

# MA3 Analysis graders
from graders.ma3_analysis.grade_analysis import grade_analysis_tab
from writers.write_ma3_analysis_results import write_ma3_analysis_results

# MA3 Visualization graders
from graders.ma3_visualization.grade_visualization import grade_visualization_tab
from writers.write_ma3_visualization_results import write_ma3_visualization_results


def _validate_ma3_sheets(workbook) -> tuple:
    """
    Validate that required MA3 sheets exist (case-insensitive).
    
    Returns:
        Tuple of (is_valid, sheet_map, missing_sheets)
        sheet_map maps canonical names to actual sheet names
    """
    required_sheets = ["Analysis", "Visualization"]
    sheet_names_lower = {s.lower(): s for s in workbook.sheetnames}
    
    sheet_map = {}
    missing_sheets = []
    
    for req in required_sheets:
        if req.lower() in sheet_names_lower:
            sheet_map[req] = sheet_names_lower[req.lower()]
        else:
            missing_sheets.append(req)
    
    is_valid = len(missing_sheets) == 0
    return is_valid, sheet_map, missing_sheets


def phase1_grade_all_students_ma3(
    submissions_path: str,
    graded_output_path: str,
    pipeline_state: Optional[Dict[str, Any]] = None
) -> None:
    """
    Grades the formula-based parts of every student's MA3 workbook.
    
    Args:
        submissions_path: Path to folder containing student submission files
        graded_output_path: Path to folder containing grading sheet templates
        pipeline_state: Optional dict for cancellation checking
    """
    logger = get_logger()
    logger.info("")
    logger.info("=" * 60)
    logger.info("PHASE 1 - Grading all MA3 students...")
    logger.info("=" * 60)

    student_files = [f for f in os.listdir(submissions_path) if f.endswith(".xlsx")]
    total_students = len(student_files)
    logger.info(f"Found {total_students} student submissions to grade")

    graded_count = 0
    error_count = 0
    skipped_count = 0

    for idx, filename in enumerate(student_files, 1):
        # Check for cancellation request
        if pipeline_state and pipeline_state.get("cancel_requested"):
            logger.warning(f"Pipeline cancelled by user after grading {graded_count} students")
            break

        # Extract student name from filename
        student_name = filename.replace("_MA3.xlsx", "").replace(".xlsx", "")
        submission_file = os.path.join(submissions_path, filename)
        grading_file = os.path.join(graded_output_path, f"{student_name}_MA3_Grade.xlsx")

        logger.info(f"[{idx}/{total_students}] Processing: {student_name}")

        student_wb = None
        grading_wb = None
        
        try:
            logger.debug(f"  Loading submission: {submission_file}")
            student_wb = load_workbook(submission_file, data_only=False)
            
            logger.debug(f"  Loading grading sheet: {grading_file}")
            grading_wb = load_workbook(grading_file)

            ws_grading = grading_wb["Grading Sheet"]

            # Validate required sheets
            is_valid, sheet_map, missing_sheets = _validate_ma3_sheets(student_wb)
            
            if missing_sheets:
                logger.warning(f"  Missing sheets for {student_name}: {missing_sheets}")

            # -----------------------------
            # ANALYSIS TAB
            # -----------------------------
            if sheet_map.get("Analysis"):
                try:
                    logger.debug(f"  Grading Analysis tab...")
                    ws_analysis = student_wb[sheet_map["Analysis"]]
                    analysis_results = grade_analysis_tab(ws_analysis, student_name)
                    write_ma3_analysis_results(ws_grading, analysis_results)
                    logger.debug(f"  Analysis tab complete")
                except Exception as e:
                    logger.warning(f"  Analysis error for {student_name}: {e}")
            else:
                logger.info(f"  Skipping Analysis (sheet missing)")
                skipped_count += 1

            # -----------------------------
            # VISUALIZATION TAB
            # -----------------------------
            if sheet_map.get("Visualization"):
                try:
                    logger.debug(f"  Grading Visualization tab...")
                    ws_viz = student_wb[sheet_map["Visualization"]]
                    viz_results = grade_visualization_tab(ws_viz)
                    write_ma3_visualization_results(ws_grading, viz_results)
                    logger.debug(f"  Visualization tab complete")
                except Exception as e:
                    logger.warning(f"  Visualization error for {student_name}: {e}")
            else:
                logger.info(f"  Skipping Visualization (sheet missing)")
                skipped_count += 1

            grading_wb.save(grading_file)
            logger.info(f"  ✓ Graded: {student_name}")
            graded_count += 1

        except Exception as e:
            logger.error(f"  ✗ Error grading {student_name}: {e}")
            error_count += 1
        
        finally:
            # Ensure workbooks are always closed
            if student_wb is not None:
                student_wb.close()
            if grading_wb is not None:
                grading_wb.close()

    # Summary
    logger.info("")
    logger.info("-" * 40)
    logger.info(f"Phase 1 Summary: {graded_count} graded, {error_count} errors, {skipped_count} sheets skipped")
    logger.info("-" * 40)
