"""
test_integration.py — Integration tests using real Excel files

These tests use actual student submission files to verify end-to-end
grading functionality. They require the sample ZIP to be present at:
~/Desktop/Major Assignment 1 - Online.zip

Tests include:
- Full pipeline execution on sample submissions
- Grading accuracy verification
- Edge case handling with real files
"""

import pytest
import sys
import os
import tempfile
import shutil
import zipfile
from pathlib import Path

# Add backend to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ============================================================
# Test Fixtures
# ============================================================

SAMPLE_ZIP_PATH = os.path.expanduser("~/Desktop/Major Assignment 1 - Online.zip")


def sample_zip_exists():
    """Check if the sample ZIP file exists."""
    return os.path.exists(SAMPLE_ZIP_PATH)


@pytest.fixture
def temp_workspace():
    """Create a temporary workspace for testing."""
    temp_dir = tempfile.mkdtemp(prefix="ma_grader_test_")
    yield temp_dir
    # Cleanup after test
    shutil.rmtree(temp_dir, ignore_errors=True)


@pytest.fixture
def extracted_submissions(temp_workspace):
    """Extract sample submissions to temp directory."""
    if not sample_zip_exists():
        pytest.skip("Sample ZIP not found - skipping integration test")
    
    extract_dir = os.path.join(temp_workspace, "extracted")
    os.makedirs(extract_dir, exist_ok=True)
    
    with zipfile.ZipFile(SAMPLE_ZIP_PATH, 'r') as zip_ref:
        zip_ref.extractall(extract_dir)
    
    return extract_dir


# ============================================================
# Integration Tests with Real Files
# ============================================================

@pytest.mark.skipif(not sample_zip_exists(), reason="Sample ZIP not found")
class TestRealFileGrading:
    """Integration tests using real student submission files."""
    
    def test_load_real_submission(self, extracted_submissions):
        """Should be able to load a real student submission file."""
        from openpyxl import load_workbook
        
        # Find first Excel file in extracted directory
        for root, dirs, files in os.walk(extracted_submissions):
            for f in files:
                if f.endswith('.xlsx') and not f.startswith('~'):
                    file_path = os.path.join(root, f)
                    
                    wb = load_workbook(file_path, data_only=False)
                    assert len(wb.sheetnames) > 0
                    wb.close()
                    return
        
        pytest.fail("No Excel files found in extracted submissions")
    
    def test_grade_real_income_analysis(self, extracted_submissions):
        """Grade Income Analysis on a real submission."""
        from openpyxl import load_workbook
        from graders.income_analysis.grade_income_analysis import grade_income_analysis
        from utilities.validate_submission import get_sheet_safe
        
        # Find first Excel file
        file_path = None
        for root, dirs, files in os.walk(extracted_submissions):
            for f in files:
                if f.endswith('.xlsx') and not f.startswith('~'):
                    file_path = os.path.join(root, f)
                    break
            if file_path:
                break
        
        if not file_path:
            pytest.skip("No Excel files found")
        
        wb = load_workbook(file_path, data_only=False)
        ws = get_sheet_safe(wb, "Income Analysis")
        
        if ws is None:
            wb.close()
            pytest.skip("Income Analysis sheet not found in this file")
        
        results = grade_income_analysis(ws)
        
        # Verify structure
        assert "name_score" in results
        assert "slope_score" in results
        assert "predictions_score" in results
        assert "scatterplot_chart_score" in results
        assert "scatterplot_trendline_score" in results
        
        # Scores should be non-negative
        assert results["name_score"] >= 0
        assert results["slope_score"] >= 0
        assert results["predictions_score"] >= 0
        
        wb.close()
    
    def test_validate_real_submission_sheets(self, extracted_submissions):
        """Validate required sheets on a real submission."""
        from openpyxl import load_workbook
        from utilities.validate_submission import validate_required_sheets
        
        files_checked = 0
        
        for root, dirs, files in os.walk(extracted_submissions):
            for f in files:
                if f.endswith('.xlsx') and not f.startswith('~'):
                    file_path = os.path.join(root, f)
                    
                    try:
                        wb = load_workbook(file_path, data_only=False)
                        is_valid, sheet_map, missing = validate_required_sheets(wb)
                        
                        # Should return proper types
                        assert isinstance(is_valid, bool)
                        assert isinstance(sheet_map, dict)
                        assert isinstance(missing, list)
                        
                        wb.close()
                        files_checked += 1
                    except Exception as e:
                        # Some files might be corrupted
                        print(f"Could not process {f}: {e}")
        
        assert files_checked > 0, "No files could be validated"
    
    def test_grade_all_submissions_in_zip(self, extracted_submissions):
        """Grade all submissions in the sample ZIP."""
        from openpyxl import load_workbook
        from graders.income_analysis.grade_income_analysis import grade_income_analysis
        from utilities.validate_submission import validate_required_sheets, get_sheet_safe
        
        graded_count = 0
        error_count = 0
        
        for root, dirs, files in os.walk(extracted_submissions):
            for f in files:
                if f.endswith('.xlsx') and not f.startswith('~'):
                    file_path = os.path.join(root, f)
                    
                    try:
                        wb = load_workbook(file_path, data_only=False)
                        is_valid, sheet_map, missing = validate_required_sheets(wb)
                        
                        if sheet_map.get("Income Analysis"):
                            ws = wb[sheet_map["Income Analysis"]]
                            results = grade_income_analysis(ws)
                            
                            # Verify grading completed
                            assert results is not None
                            graded_count += 1
                        
                        wb.close()
                    except Exception as e:
                        error_count += 1
                        print(f"Error processing {f}: {e}")
        
        print(f"Graded {graded_count} files, {error_count} errors")
        assert graded_count > 0, "No files were successfully graded"


# ============================================================
# Test Edge Cases with Real Files
# ============================================================

class TestEdgeCasesWithMockFiles:
    """Test edge cases using mock/synthetic files."""
    
    def test_empty_excel_file(self, temp_workspace):
        """Should handle an empty Excel file gracefully."""
        from openpyxl import Workbook
        from utilities.validate_submission import validate_required_sheets
        
        # Create empty workbook
        wb = Workbook()
        file_path = os.path.join(temp_workspace, "empty.xlsx")
        wb.save(file_path)
        
        # Validate
        is_valid, sheet_map, missing = validate_required_sheets(wb)
        
        assert is_valid is False
        assert len(missing) == 3  # All required sheets missing
        
        wb.close()
    
    def test_workbook_with_wrong_sheets(self, temp_workspace):
        """Should handle workbook with wrong sheet names."""
        from openpyxl import Workbook
        from utilities.validate_submission import validate_required_sheets
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Wrong Sheet Name"
        wb.create_sheet("Another Wrong Sheet")
        
        is_valid, sheet_map, missing = validate_required_sheets(wb)
        
        assert is_valid is False
        assert "Income Analysis" in missing
        
        wb.close()
    
    def test_workbook_with_partial_sheets(self, temp_workspace):
        """Should report only missing sheets."""
        from openpyxl import Workbook
        from utilities.validate_submission import validate_required_sheets
        
        wb = Workbook()
        wb.active.title = "Income Analysis"
        wb.create_sheet("Unit Conversions")
        # Missing: Currency Conversion
        
        is_valid, sheet_map, missing = validate_required_sheets(wb)
        
        assert is_valid is False
        assert len(missing) == 1
        assert "Currency Conversion" in missing
        assert sheet_map["Income Analysis"] is not None
        
        wb.close()


# ============================================================
# Test Corrupted File Handling
# ============================================================

class TestCorruptedFileHandling:
    """Test handling of corrupted or invalid files."""
    
    def test_invalid_file_extension(self, temp_workspace):
        """Should handle file with wrong extension."""
        # Create a text file with .xlsx extension
        file_path = os.path.join(temp_workspace, "fake.xlsx")
        with open(file_path, 'w') as f:
            f.write("This is not an Excel file")
        
        from openpyxl import load_workbook
        
        with pytest.raises(Exception):
            load_workbook(file_path)
    
    def test_truncated_excel_file(self, temp_workspace):
        """Should handle truncated Excel file."""
        from openpyxl import Workbook
        
        # Create valid workbook
        wb = Workbook()
        file_path = os.path.join(temp_workspace, "truncated.xlsx")
        wb.save(file_path)
        wb.close()
        
        # Truncate the file
        with open(file_path, 'r+b') as f:
            f.seek(100)
            f.truncate()
        
        from openpyxl import load_workbook
        
        with pytest.raises(Exception):
            load_workbook(file_path)
    
    def test_password_protected_handling(self):
        """Document behavior for password-protected files."""
        # Password-protected files would raise an exception
        # This test documents expected behavior
        pass


# ============================================================
# Test ZIP Processing
# ============================================================

@pytest.mark.skipif(not sample_zip_exists(), reason="Sample ZIP not found")
class TestZIPProcessing:
    """Tests for ZIP file handling."""
    
    def test_zip_extraction(self, temp_workspace):
        """Should extract ZIP file correctly."""
        extract_dir = os.path.join(temp_workspace, "extracted")
        
        with zipfile.ZipFile(SAMPLE_ZIP_PATH, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
        
        # Verify extraction
        assert os.path.exists(extract_dir)
        
        # Count Excel files
        xlsx_count = 0
        for root, dirs, files in os.walk(extract_dir):
            xlsx_count += len([f for f in files if f.endswith('.xlsx')])
        
        assert xlsx_count > 0, "No Excel files extracted"
    
    def test_nested_folder_structure(self, temp_workspace):
        """ZIP should have expected folder structure."""
        extract_dir = os.path.join(temp_workspace, "extracted")
        
        with zipfile.ZipFile(SAMPLE_ZIP_PATH, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
        
        # Check for student folders (folders containing student names)
        student_folders = []
        for item in os.listdir(extract_dir):
            item_path = os.path.join(extract_dir, item)
            if os.path.isdir(item_path):
                for sub_item in os.listdir(item_path):
                    sub_path = os.path.join(item_path, sub_item)
                    if os.path.isdir(sub_path):
                        student_folders.append(sub_item)
        
        assert len(student_folders) > 0, "No student folders found"
