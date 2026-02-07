"""
test_ma3_visualization.py — Unit tests for MA3 Visualization tab graders

Tests cover:
- Bin table validation (Min, Max, Width)
- Frequency distribution limits (Lower, Upper)
- Frequency distribution values (Title, Frequency, Relative Frequency)
- Histogram validation (chart type, data, labels, titles)
- Formatting checks
"""

import pytest
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from graders.ma3_visualization.check_bin_table import (
    _check_min_formula, _check_max_formula, _check_width_formula
)
from graders.ma3_visualization.check_freq_dist import (
    _check_lower_limit_formula, _check_upper_limit_formula,
    _check_title_formula, _check_frequency_formula, _check_relative_freq_formula
)
from graders.ma3_visualization.check_histogram import _extract_title_text


# ============================================================
# Bin Table Tests
# ============================================================

class TestBinMinFormula:
    """Tests for Bin Min formula validation."""
    
    def test_basic_min(self):
        """Basic MIN formula."""
        assert _check_min_formula("=MIN(B12:B61)") is True
    
    def test_min_with_dollars(self):
        """MIN with absolute references."""
        assert _check_min_formula("=MIN($B$12:$B$61)") is True
    
    def test_min_with_offset(self):
        """MIN with slight offset (B14 start)."""
        assert _check_min_formula("=MIN(B14:B63)") is True
    
    def test_not_min(self):
        """Non-MIN function should fail."""
        assert _check_min_formula("=MAX(B12:B61)") is False
    
    def test_min_wrong_column(self):
        """MIN with wrong column should fail."""
        assert _check_min_formula("=MIN(A12:A61)") is False


class TestBinMaxFormula:
    """Tests for Bin Max formula validation."""
    
    def test_basic_max(self):
        """Basic MAX formula."""
        assert _check_max_formula("=MAX(B12:B61)") is True
    
    def test_max_with_dollars(self):
        """MAX with absolute references."""
        assert _check_max_formula("=MAX($B$12:$B$61)") is True
    
    def test_not_max(self):
        """Non-MAX function should fail."""
        assert _check_max_formula("=MIN(B12:B61)") is False


class TestBinWidthFormula:
    """Tests for Bin Width formula validation."""
    
    def test_basic_width(self):
        """Basic width formula."""
        assert _check_width_formula("=(E23-E22)/10") is True
    
    def test_width_with_dollars(self):
        """Width with absolute references."""
        assert _check_width_formula("=($E$23-$E$22)/11") is True
    
    def test_width_different_divisor(self):
        """Width with different number of bins."""
        assert _check_width_formula("=(E23-E22)/12") is True
    
    def test_width_with_max_min(self):
        """Width calculated directly with MAX/MIN."""
        assert _check_width_formula("=(MAX(B12:B61)-MIN(B12:B61))/10") is True
    
    def test_width_no_division(self):
        """Formula without division should fail."""
        assert _check_width_formula("=E23-E22") is False


# ============================================================
# Frequency Distribution Limit Tests
# ============================================================

class TestLowerLimitFormula:
    """Tests for Lower Limit formula validation."""
    
    def test_first_row_e22(self):
        """First row should reference E22 (bin min)."""
        assert _check_lower_limit_formula("=E22", 28) is True
    
    def test_first_row_with_offset(self):
        """First row with subtraction offset."""
        assert _check_lower_limit_formula("=E22-0.1", 28) is True
    
    def test_subsequent_row(self):
        """Subsequent rows should reference previous upper."""
        assert _check_lower_limit_formula("=E28", 29) is True
    
    def test_subsequent_row_30(self):
        """Row 30 should reference E29."""
        assert _check_lower_limit_formula("=E29", 30) is True
    
    def test_wrong_reference(self):
        """Wrong cell reference should fail."""
        assert _check_lower_limit_formula("=E20", 28) is False


class TestUpperLimitFormula:
    """Tests for Upper Limit formula validation."""
    
    def test_basic_upper(self):
        """Basic upper limit formula."""
        assert _check_upper_limit_formula("=D28+$E$24", 28) is True
    
    def test_upper_without_dollars(self):
        """Upper limit without absolute reference."""
        assert _check_upper_limit_formula("=D28+E24", 28) is True
    
    def test_upper_row_35(self):
        """Upper limit for row 35."""
        assert _check_upper_limit_formula("=D35+$E$24", 35) is True
    
    def test_wrong_row(self):
        """Wrong row reference should fail."""
        assert _check_upper_limit_formula("=D27+$E$24", 28) is False


# ============================================================
# Frequency Distribution Value Tests
# ============================================================

class TestTitleFormula:
    """Tests for Title of Bin (midpoint) formula validation."""
    
    def test_basic_midpoint(self):
        """Basic midpoint formula."""
        assert _check_title_formula("=(D28+E28)/2", 28) is True
    
    def test_midpoint_with_dollars(self):
        """Midpoint with absolute references."""
        assert _check_title_formula("=($D$28+$E$28)/2", 28) is True
    
    def test_midpoint_row_35(self):
        """Midpoint for row 35."""
        assert _check_title_formula("=(D35+E35)/2", 35) is True
    
    def test_not_midpoint(self):
        """Non-midpoint formula should fail."""
        assert _check_title_formula("=D28", 28) is False


class TestFrequencyFormula:
    """Tests for Frequency formula validation."""
    
    def test_countifs(self):
        """COUNTIFS formula."""
        assert _check_frequency_formula('=COUNTIFS($B$12:$B$61,">="&D28,$B$12:$B$61,"<="&E28)') is True
    
    def test_countif(self):
        """COUNTIF formula."""
        assert _check_frequency_formula('=COUNTIF(B12:B61,">="&D28)') is True
    
    def test_frequency_function(self):
        """FREQUENCY function."""
        assert _check_frequency_formula("=FREQUENCY(B12:B61,E28:E38)") is True
    
    def test_sumproduct(self):
        """SUMPRODUCT formula."""
        assert _check_frequency_formula("=SUMPRODUCT((B12:B61>=D28)*(B12:B61<=E28))") is True
    
    def test_not_counting(self):
        """Non-counting formula should fail."""
        assert _check_frequency_formula("=SUM(B12:B61)") is False


class TestRelativeFreqFormula:
    """Tests for Relative Frequency formula validation."""
    
    def test_basic_relative(self):
        """Basic relative frequency formula."""
        assert _check_relative_freq_formula("=G28/50", 28) is True
    
    def test_relative_with_sum(self):
        """Relative frequency with SUM."""
        assert _check_relative_freq_formula("=G28/SUM($G$28:$G$38)", 28) is True
    
    def test_relative_row_35(self):
        """Relative frequency for row 35."""
        assert _check_relative_freq_formula("=G35/50", 35) is True
    
    def test_missing_freq_ref(self):
        """Formula without frequency reference should fail."""
        assert _check_relative_freq_formula("=50/100", 28) is False
    
    def test_no_division(self):
        """Formula without division should fail."""
        assert _check_relative_freq_formula("=G28", 28) is False


# ============================================================
# Histogram Tests
# ============================================================

class TestHistogramTitleExtraction:
    """Tests for chart title extraction."""
    
    def test_none_title(self):
        """None title should return None."""
        assert _extract_title_text(None) is None


# ============================================================
# Integration Tests with Real Workbook
# ============================================================

class TestRealVisualization:
    """Integration tests using actual student submission."""
    
    @pytest.fixture
    def student_workbook(self):
        """Load a real student workbook for testing."""
        from openpyxl import load_workbook
        
        test_path = "/tmp/ma3-test/Major Assignment 3 - Online/Amber_Carbonneau_21276491"
        if not os.path.exists(test_path):
            pytest.skip("Test data not available")
        
        xlsx_files = [f for f in os.listdir(test_path) if f.endswith('.xlsx')]
        if not xlsx_files:
            pytest.skip("No xlsx file found")
        
        wb = load_workbook(os.path.join(test_path, xlsx_files[0]), data_only=False)
        yield wb
        wb.close()
    
    def test_full_visualization_grading(self, student_workbook):
        """Test complete Visualization tab grading."""
        from graders.ma3_visualization import grade_visualization_tab
        
        ws = student_workbook["Visualization"]
        results = grade_visualization_tab(ws)
        
        # Amber should have high scores (95/100 total, 46/46 viz)
        assert results["bin_score"] == 6.0
        assert results["limits_score"] == 12.0
        assert results["freqdist_score"] >= 17.0  # Near perfect
    
    def test_histogram_detection(self, student_workbook):
        """Test that histogram is detected correctly."""
        from graders.ma3_visualization.check_histogram import check_histogram
        
        ws = student_workbook["Visualization"]
        score, feedback = check_histogram(ws)
        
        # Should find a histogram
        assert score >= 4.0  # At least chart found + some criteria


class TestLowScoringStudent:
    """Test with a student who has lower scores."""
    
    @pytest.fixture
    def low_score_workbook(self):
        """Load a lower-scoring student workbook."""
        from openpyxl import load_workbook
        
        test_path = "/tmp/ma3-test/Major Assignment 3 - Online/Nicole_Smith_21280001"
        if not os.path.exists(test_path):
            pytest.skip("Test data not available")
        
        xlsx_files = [f for f in os.listdir(test_path) if f.endswith('.xlsx')]
        if not xlsx_files:
            pytest.skip("No xlsx file found")
        
        wb = load_workbook(os.path.join(test_path, xlsx_files[0]), data_only=False)
        yield wb
        wb.close()
    
    def test_detects_issues(self, low_score_workbook):
        """Test that grader correctly identifies issues."""
        from graders.ma3_visualization import grade_visualization_tab
        
        ws = low_score_workbook["Visualization"]
        results = grade_visualization_tab(ws)
        
        # Nicole has some issues, should not be perfect
        total = sum(v for k, v in results.items() 
                   if 'score' in k and isinstance(v, (int, float)))
        
        assert total < 46.0  # Not perfect score
        assert total > 20.0  # But not zero either


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
