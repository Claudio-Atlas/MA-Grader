# utilities/validate_submission.py

"""
Validates student submission workbooks before grading.

Ensures required sheets exist (with case-insensitive matching) to prevent
KeyError crashes when students rename or delete sheets.
"""

from typing import List, Tuple, Dict, Optional
from openpyxl import Workbook

from utilities.logger import get_logger


# Required sheets for MA1 grading
REQUIRED_SHEETS = [
    "Income Analysis",
    "Unit Conversions",
    "Currency Conversion"
]


def validate_required_sheets(
    workbook: Workbook,
    required_sheets: List[str] = None
) -> Tuple[bool, Dict[str, Optional[str]], List[str]]:
    """
    Validates that a workbook contains all required sheets.
    
    Uses case-insensitive matching to handle common student mistakes
    (e.g., "income analysis" instead of "Income Analysis").
    
    Args:
        workbook: An openpyxl Workbook object
        required_sheets: List of required sheet names (defaults to REQUIRED_SHEETS)
    
    Returns:
        Tuple of:
        - is_valid: bool - True if all required sheets are present
        - sheet_map: Dict[str, Optional[str]] - Maps required name -> actual name found (or None)
        - missing: List[str] - List of missing sheet names
    
    Example:
        is_valid, sheet_map, missing = validate_required_sheets(wb)
        if not is_valid:
            logger.warning(f"Missing sheets: {missing}")
        else:
            # Use sheet_map to get actual sheet names
            income_ws = wb[sheet_map["Income Analysis"]]
    """
    logger = get_logger()
    
    if required_sheets is None:
        required_sheets = REQUIRED_SHEETS
    
    # Get actual sheet names from workbook
    actual_sheets = workbook.sheetnames
    
    # Build lowercase lookup for case-insensitive matching
    actual_sheets_lower = {name.lower().strip(): name for name in actual_sheets}
    
    sheet_map = {}
    missing = []
    
    for required in required_sheets:
        required_lower = required.lower().strip()
        
        if required_lower in actual_sheets_lower:
            # Found with case-insensitive match
            sheet_map[required] = actual_sheets_lower[required_lower]
        else:
            # Not found
            sheet_map[required] = None
            missing.append(required)
    
    is_valid = len(missing) == 0
    
    if missing:
        logger.debug(f"Missing sheets in workbook: {missing}")
    
    return is_valid, sheet_map, missing


def get_sheet_safe(
    workbook: Workbook,
    sheet_name: str,
    case_insensitive: bool = True
) -> Optional[object]:
    """
    Safely gets a worksheet by name, with optional case-insensitive matching.
    
    Args:
        workbook: An openpyxl Workbook object
        sheet_name: The sheet name to find
        case_insensitive: If True, matches regardless of case
    
    Returns:
        The worksheet if found, None otherwise
    """
    if not case_insensitive:
        return workbook[sheet_name] if sheet_name in workbook.sheetnames else None
    
    sheet_name_lower = sheet_name.lower().strip()
    
    for actual_name in workbook.sheetnames:
        if actual_name.lower().strip() == sheet_name_lower:
            return workbook[actual_name]
    
    return None


def log_missing_sheets(student_name: str, missing_sheets: List[str]) -> None:
    """
    Logs missing sheets for a student submission.
    
    Args:
        student_name: The student's name/identifier
        missing_sheets: List of missing sheet names
    """
    logger = get_logger()
    sheets_str = ", ".join(missing_sheets)
    logger.warning(f"{student_name}: Missing required sheets: {sheets_str}")
